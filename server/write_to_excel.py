import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle,Font,Border
from openpyxl.worksheet.table import Table,TableStyleInfo

def create_new_excel_file(name):
    wb = Workbook()
    file_name = os.path.join(os.path.dirname(__file__), f"{name}.xlsx")
    
    # exists = os.path.exists(file_name)
    # num = 1
    # while exists: 
    #     file_name = os.path.join(os.path.dirname(__file__), f"{name}({num}).xlsx")
    #     num +=1
    #     exists = os.path.exists(file_name)
    
    return file_name,wb

def add_sheet_title(ws):

   ws.title = "Workout"

def add_title(ws,title):
    font = Font(name='Ariel', size=18,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='00FF0000')
    cell = ws['A1']
    cell.value = title
    cell.font = font
    merge_cells(ws,"A1","D1")

def merge_cells(ws, start_cell,end_cell):
    cell_range = f"{start_cell}:{end_cell}"
    ws.merge_cells(cell_range)
    
def add_excercise(ws):
    data = [
        ["Rowing","","","","","","","","1",],
        ["Squats","","","","","","","","1",]
    ]
    ws.append(["","Set1","","Set2","","Set3",""])
    ws.append(["Exercise","Reps","Wgt","Reps","Wgt","Reps","Wgt"])
    for row in data:
        ws.append(row)
    start_row = 1
    for row in ws.iter_rows():
        print(row)
        if set(row) == None:
            break
        start_row += 1
    print(start_row)
            
        
    
        
    
        
    #tab = Table(displayName="Day", ref="")
        
    
    
       
    
    
    
    
    

file_name, workbook = create_new_excel_file("workout")
#workbook = load_workbook(file_name)#os.path.join(os.path.dirname(__file__),"workout.xlsx"))
ws = workbook.active
add_sheet_title(ws)
add_title(ws, "Programme Name")
add_excercise(ws)
workbook.save(file_name)
